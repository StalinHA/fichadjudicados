import streamlit as st
import pandas as pd
import json
import requests
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
import base64
from io import BytesIO
import zipfile
import io
import re
from difflib import SequenceMatcher
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============ CONFIGURACIÓN PARA HUGGING FACE SPACES ============
os.environ['STREAMLIT_STATIC_URL_PREFIX'] = ''

# Configuración de la página
st.set_page_config(
    page_title="Dashboard de Análisis de Productos",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============ LISTA COMPLETA DE MARCAS ============
MARCAS_COMPLETAS = [
    'TRIUMPH BOARD', 'VASTEC', 'RHINOBOX', 'LENOVO', 'EXIN', 'M4X', 'KENYA TECHNOLOGY',
    'HP', 'MADI-TEK', 'INVESTMENT & BUSINESS SMART SBI', 'ADVANCE', 'QUI-TECH', 'TEXCOPER',
    'WIDETEK', 'IQTOUCH', 'LG', 'ONESCREEN', 'HUAWEI', 'INOTEC', 'DELL', 'I3', 'ASUS',
    'QUAMTU', 'KODAK', 'RICOH', 'SHARP', 'CIBER', 'HAO TECH', 'BROTHER', 'VIEWSONIC',
    'AVISION', 'SAMSUNG', 'ALLWIYA', 'GAMEMAX', 'DYNABOOK', 'HIPPOBOX', 'CONTEX', 'INNEX',
    'CTOUCH', 'HIKVISION', 'ZKT ECO', 'YEALINK', 'TEROS', 'SILVER VOLT', 'QOSOFT',
    'MIMIO', 'HAITECH', 'OPTOMA TECHNOLOGY INC', 'GROWTH HACK', 'MSI', 'XEROX', 'QOMO',
    'EPSON', 'CLEVERTOUCH', 'I2S INNOVATIVE IMAGING SOLUTIONS', 'IQ BOARD', 'GCS', 'COLORTRAC',
    'CANON', 'BOOKEYE', 'JFA TECHNOLOGY', 'AMC', 'MAXTIC', 'SANDISK', 'KINGSTON', 'ADATA', 'NEW KRAL'
]

# ============ CATEGORIAS OFICIALES ============
CATEGORIAS_OFICIALES = {
    '11743': 'COMPUTADORA PORTATIL',
    '11744': 'ESTACION DE TRABAJO PORTATIL',
    '11745': 'TABLETA',
    '11738': 'ESCANER DE DOCUMENTOS',
    '11735': 'COMPUTADORA DE ESCRITORIO',
    '11736': 'COMPUTADORA TODO EN UNO',
    '11740': 'ESTACION DE TRABAJO',
    '11741': 'MONITOR',
    '11742': 'PANTALLA PUBLICITARIA',
    '11747': 'DISPOSITIVOS DE ALMACENAMIENTO EXTERNO',
    '11749': 'PANTALLA INTERACTIVA',
    '11751': 'DISPOSITIVOS DE ALMACENAMIENTO INTERNO'
}

CATEGORIAS_NOMBRES = list(CATEGORIAS_OFICIALES.values())

# Estilos CSS
st.markdown("""
<style>
    .main-header {font-size: 2.5rem; color: #1f77b4 !important; text-align: center; padding: 1rem; background: linear-gradient(90deg, #f0f2f6, #ffffff); border-radius: 10px; margin-bottom: 2rem;}
    .metric-card {background: white; padding: 1.5rem; border-radius: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); text-align: center; margin: 0.5rem 0; transition: transform 0.3s;}
    .metric-card:hover {transform: translateY(-5px); box-shadow: 0 4px 8px rgba(0,0,0,0.2);}
    .metric-value {font-size: 2.5rem; font-weight: bold; color: #1f77b4;}
    .metric-label {font-size: 0.9rem; color: #333; margin-top: 0.5rem;}
    .metric-sub {font-size: 0.8rem; color: #666; margin-top: 0.2rem;}
    .stTabs [data-baseweb="tab"] {height: 50px; background-color: #f0f2f6; border-radius: 5px; padding: 10px 20px; font-weight: bold; color: #333;}
    .stTabs [aria-selected="true"] {background-color: #1f77b4; color: white !important;}
    .resumen-activo-header {background: linear-gradient(135deg, #28a745 0%, #20c997 100%); padding: 1.5rem; border-radius: 15px; margin: 1rem 0; text-align: center;}
    .resumen-activo-header h2 {color: white !important; margin: 0;}
    .resumen-activo-header p {color: rgba(255,255,255,0.9) !important; margin: 0;}
    .analisis-header {background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); padding: 1rem; border-radius: 10px; margin: 1rem 0; text-align: center;}
    .analisis-header h2 {color: white !important; margin: 0;}
    .search-container {background: #f8f9fa; padding: 1.5rem; border-radius: 10px; margin: 1rem 0;}
    .result-header {background: #d4edda; padding: 1rem; border-radius: 10px; margin: 1rem 0; border-left: 4px solid #28a745;}
    .no-result {background: #f8d7da; padding: 1rem; border-radius: 10px; margin: 1rem 0; border-left: 4px solid #dc3545;}
    .download-btn {background-color: #28a745; color: white; padding: 10px 20px; border-radius: 5px; text-decoration: none; display: inline-block; font-weight: bold;}
    .download-btn:hover {background-color: #218838; color: white;}
</style>
""", unsafe_allow_html=True)

# ============ FUNCIÓN PARA CREAR EXCEL CON FORMATO ============
def crear_excel_formateado(df, nombre_archivo="reporte_junio_2026.xlsx"):
    """Crea un archivo Excel con formato y colores"""
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir datos
        df.to_excel(writer, sheet_name='Junio 2026', index=False)
        
        # Obtener el workbook y la hoja
        workbook = writer.book
        worksheet = writer.sheets['Junio 2026']
        
        # ============ DEFINIR ESTILOS ============
        # Colores
        verde_claro = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        verde_oscuro = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
        rojo_claro = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        rojo_oscuro = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
        amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        azul_claro = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        morado = PatternFill(start_color="D5C6E0", end_color="D5C6E0", fill_type="solid")
        
        # Fuentes
        fuente_blanco = Font(color="FFFFFF", bold=True)
        fuente_negro = Font(color="000000", bold=True)
        
        # Bordes
        borde_fino = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alineación
        centrado = Alignment(horizontal='center', vertical='center')
        izquierda = Alignment(horizontal='left', vertical='center')
        derecha = Alignment(horizontal='right', vertical='center')
        
        # ============ APLICAR ESTILOS A LA TABLA ============
        # Obtener el rango de datos
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Obtener nombres de columnas
        col_estado_ficha = None
        col_estado_oferta = None
        col_precio = None
        col_marca = None
        col_categoria = None
        
        for col in range(1, max_col + 1):
            header = worksheet.cell(row=1, column=col).value
            if header == 'estado_ficha':
                col_estado_ficha = col
            elif header == 'estado_oferta':
                col_estado_oferta = col
            elif header == 'precio':
                col_precio = col
            elif header == 'marca':
                col_marca = col
            elif header == 'categoria':
                col_categoria = col
        
        # ============ 1. HEADER (Fila 1) ============
        for col in range(1, max_col + 1):
            celda = worksheet.cell(row=1, column=col)
            celda.fill = azul_claro
            celda.font = fuente_negro
            celda.alignment = centrado
            celda.border = borde_fino
        
        # ============ 2. COLORES POR ESTADO ============
        if col_estado_ficha and col_estado_oferta:
            for row in range(2, max_row + 1):
                estado_ficha = worksheet.cell(row=row, column=col_estado_ficha).value
                estado_oferta = worksheet.cell(row=row, column=col_estado_oferta).value
                
                # Determinar color según estado
                if estado_ficha == 'OFERTADA' and estado_oferta == 'VIGENTE':
                    fill = verde_claro
                    font = Font(color="006100")
                elif estado_ficha == 'OFERTADA' and estado_oferta != 'VIGENTE':
                    fill = amarillo
                    font = Font(color="9C5700")
                elif estado_ficha == 'EXCLUIDA':
                    fill = rojo_claro
                    font = Font(color="9C0006")
                else:
                    fill = gris
                    font = Font(color="333333")
                
                # Aplicar a toda la fila
                for col in range(1, max_col + 1):
                    celda = worksheet.cell(row=row, column=col)
                    celda.fill = fill
                    celda.font = font
                    celda.border = borde_fino
                    celda.alignment = izquierda if col not in [col_precio] else derecha
        
        # ============ 3. FORMATO DE PRECIOS ============
        if col_precio:
            for row in range(2, max_row + 1):
                celda = worksheet.cell(row=row, column=col_precio)
                try:
                    valor = float(celda.value)
                    celda.value = f"${valor:,.2f}"
                    celda.alignment = derecha
                except:
                    pass
        
        # ============ 4. AJUSTAR ANCHO DE COLUMNAS ============
        for col in range(1, max_col + 1):
            max_length = 0
            for row in range(1, max_row + 1):
                valor = worksheet.cell(row=row, column=col).value
                if valor:
                    max_length = max(max_length, len(str(valor)))
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[chr(64 + col)].width = adjusted_width
        
        # ============ 5. CONGELAR PANEL ============
        worksheet.freeze_panes = 'A2'
        
        # ============ 6. AGREGAR HOJA DE RESUMEN ============
        # Crear hoja de resumen
        resumen = workbook.create_sheet("Resumen")
        
        # Título
        resumen['A1'] = "RESUMEN JUNIO 2026"
        resumen['A1'].font = Font(size=16, bold=True)
        resumen.merge_cells('A1:B1')
        
        # Estadísticas
        resumen['A3'] = "Métrica"
        resumen['B3'] = "Valor"
        resumen['A3'].fill = azul_claro
        resumen['B3'].fill = azul_claro
        resumen['A3'].font = fuente_negro
        resumen['B3'].font = fuente_negro
        
        fila_resumen = 4
        metricas = [
            ("Total de fichas", len(df)),
            ("Marcas diferentes", df['marca'].nunique() if 'marca' in df.columns else 0),
            ("Categorías diferentes", df['categoria'].nunique() if 'categoria' in df.columns else 0),
            ("Activas (OFERTADA + VIGENTE)", df[(df['estado_ficha'] == 'OFERTADA') & (df['estado_oferta'] == 'VIGENTE')].shape[0] if 'estado_ficha' in df.columns and 'estado_oferta' in df.columns else 0),
            ("Precio Promedio", f"${df['precio_float'].mean():,.2f}" if 'precio_float' in df.columns else "$0"),
            ("Precio Mínimo", f"${df['precio_float'].min():,.2f}" if 'precio_float' in df.columns else "$0"),
            ("Precio Máximo", f"${df['precio_float'].max():,.2f}" if 'precio_float' in df.columns else "$0")
        ]
        
        for metric, value in metricas:
            resumen[f'A{fila_resumen}'] = metric
            resumen[f'B{fila_resumen}'] = value
            fila_resumen += 1
        
        # Ajustar columnas de resumen
        resumen.column_dimensions['A'].width = 30
        resumen.column_dimensions['B'].width = 20
        
        # ============ 7. HOJA DE MARCAS ============
        if 'marca' in df.columns:
            hoja_marcas = workbook.create_sheet("Marcas")
            df_marcas = df['marca'].value_counts().reset_index()
            df_marcas.columns = ['Marca', 'Cantidad']
            
            # Escribir datos
            for r in dataframe_to_rows(df_marcas, index=False, header=True):
                hoja_marcas.append(r)
            
            # Estilos
            for col in range(1, 3):
                hoja_marcas.cell(row=1, column=col).fill = azul_claro
                hoja_marcas.cell(row=1, column=col).font = fuente_negro
                hoja_marcas.cell(row=1, column=col).alignment = centrado
            
            for row in range(2, len(df_marcas) + 2):
                for col in range(1, 3):
                    hoja_marcas.cell(row=row, column=col).border = borde_fino
                    if col == 1:
                        hoja_marcas.cell(row=row, column=col).alignment = izquierda
                    else:
                        hoja_marcas.cell(row=row, column=col).alignment = centrado
            
            hoja_marcas.column_dimensions['A'].width = 25
            hoja_marcas.column_dimensions['B'].width = 15
        
        # ============ 8. HOJA DE CATEGORÍAS ============
        if 'categoria' in df.columns:
            hoja_categorias = workbook.create_sheet("Categorías")
            df_categorias = df['categoria'].value_counts().reset_index()
            df_categorias.columns = ['Categoría', 'Cantidad']
            
            for r in dataframe_to_rows(df_categorias, index=False, header=True):
                hoja_categorias.append(r)
            
            for col in range(1, 3):
                hoja_categorias.cell(row=1, column=col).fill = azul_claro
                hoja_categorias.cell(row=1, column=col).font = fuente_negro
                hoja_categorias.cell(row=1, column=col).alignment = centrado
            
            for row in range(2, len(df_categorias) + 2):
                for col in range(1, 3):
                    hoja_categorias.cell(row=row, column=col).border = borde_fino
                    if col == 1:
                        hoja_categorias.cell(row=row, column=col).alignment = izquierda
                    else:
                        hoja_categorias.cell(row=row, column=col).alignment = centrado
            
            hoja_categorias.column_dimensions['A'].width = 30
            hoja_categorias.column_dimensions['B'].width = 15
    
    output.seek(0)
    return output

# ============ FUNCIONES DE UTILIDAD ============
def convertir_fecha(fecha_str):
    if not isinstance(fecha_str, str):
        return pd.NaT
    try:
        fecha_limpia = fecha_str.replace('a. m.', 'AM').replace('p. m.', 'PM')
        return pd.to_datetime(fecha_limpia, format='%d/%m/%Y %I:%M:%S %p', errors='coerce')
    except:
        return pd.NaT

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = texto.upper().strip()
    texto = re.sub(r'[^A-Z0-9]', '', texto)
    return texto

def extraer_marca(descripcion):
    if not isinstance(descripcion, str):
        return "Desconocida"
    desc_upper = descripcion.upper()
    for marca in sorted(MARCAS_COMPLETAS, key=len, reverse=True):
        if marca.upper() in desc_upper:
            return marca
    if "HP" in desc_upper:
        return "HP"
    elif "LENOVO" in desc_upper:
        return "LENOVO"
    elif "DELL" in desc_upper:
        return "DELL"
    elif "ASUS" in desc_upper:
        return "ASUS"
    elif "SAMSUNG" in desc_upper:
        return "SAMSUNG"
    elif "LG" in desc_upper:
        return "LG"
    return "Otra"

def extraer_categoria(descripcion):
    if not isinstance(descripcion, str):
        return "Sin Categoria"
    desc_upper = descripcion.upper()
    for nombre in CATEGORIAS_NOMBRES:
        if nombre.upper() in desc_upper:
            return nombre
    if "PORTATIL" in desc_upper or "NOTEBOOK" in desc_upper or "LAPTOP" in desc_upper:
        return "COMPUTADORA PORTATIL"
    elif "ESCRITORIO" in desc_upper or "DESKTOP" in desc_upper or "TORRE" in desc_upper:
        return "COMPUTADORA DE ESCRITORIO"
    elif "SERVIDOR" in desc_upper or "SERVER" in desc_upper:
        return "ESTACION DE TRABAJO"
    elif "MONITOR" in desc_upper or "PANTALLA" in desc_upper:
        return "MONITOR" if "INTERACTIVA" not in desc_upper else "PANTALLA INTERACTIVA"
    elif "TABLETA" in desc_upper:
        return "TABLETA"
    elif "ESCANER" in desc_upper:
        return "ESCANER DE DOCUMENTOS"
    return "Otro"

# ============ FUNCIONES DE BÚSQUEDA ============
def buscar_numero_parte(df, numero_parte, umbral=40):
    if df is None or df.empty or not numero_parte:
        return None
    
    numero_parte_original = numero_parte.strip()
    numero_parte_normalizado = normalizar_texto(numero_parte_original)
    resultados = []
    
    for idx, row in df.iterrows():
        descripcion = str(row['descripcion'])
        descripcion_normalizada = normalizar_texto(descripcion)
        coincidencia = 0
        
        if numero_parte_normalizado and numero_parte_normalizado in descripcion_normalizada:
            coincidencia = 100
        else:
            partes = numero_parte_original.split()
            for parte in partes:
                parte_norm = normalizar_texto(parte)
                if len(parte_norm) >= 3 and parte_norm in descripcion_normalizada:
                    coincidencia_parcial = (len(parte_norm) / len(numero_parte_normalizado)) * 100
                    coincidencia = max(coincidencia, min(100, coincidencia_parcial))
            if coincidencia < 30:
                matcher = SequenceMatcher(None, numero_parte_normalizado, descripcion_normalizada)
                coincidencia = max(coincidencia, matcher.ratio() * 100)
            if coincidencia < 40 and len(numero_parte_normalizado) >= 4:
                for i in range(len(numero_parte_normalizado) - 3):
                    subcadena = numero_parte_normalizado[i:i+4]
                    if subcadena in descripcion_normalizada:
                        coincidencia = max(coincidencia, 50)
                        break
        
        if coincidencia >= umbral:
            resultados.append({
                'ID': row['ID_ProductoOfertado'],
                'descripcion': row['descripcion'],
                'marca': row.get('marca', 'Desconocida'),
                'categoria': row.get('categoria', 'Desconocida'),
                'precio': row.get('precio', '0'),
                'estado_ficha': row.get('estado_ficha', ''),
                'estado_oferta': row.get('estado_oferta', ''),
                'fecha_publicacion': row.get('fecha_publicacion', ''),
                'fecha_adjudicacion': row.get('fecha_adjudicacion', ''),
                'coincidencia': round(coincidencia, 1),
                'es_activo': row.get('es_activo', False),
                'es_junio_2026': row.get('es_junio_2026', False)
            })
    
    if not resultados:
        return None
    resultados_df = pd.DataFrame(resultados)
    resultados_df = resultados_df.sort_values('coincidencia', ascending=False)
    return resultados_df

# ============ CACHING DE DATOS ============
@st.cache_data(ttl=3600, show_spinner=False)
def cargar_datos_github(repo_owner, repo_name, branch="main", folder=""):
    records = []
    
    if folder:
        api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{folder}"
    else:
        api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents"
    
    try:
        response = requests.get(api_url)
        if response.status_code != 200:
            return None, f"Error al acceder al repositorio: {response.status_code}"
        
        files = response.json()
        archivos_encontrados = []
        
        for file in files:
            if file['type'] == 'file':
                nombre = file['name'].lower()
                if nombre.endswith('.json') or nombre.endswith('.zip'):
                    archivos_encontrados.append(file)
            elif file['type'] == 'dir':
                sub_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{file['path']}"
                sub_response = requests.get(sub_url)
                if sub_response.status_code == 200:
                    sub_files = sub_response.json()
                    for sub_file in sub_files:
                        if sub_file['type'] == 'file':
                            nombre = sub_file['name'].lower()
                            if nombre.endswith('.json') or nombre.endswith('.zip'):
                                archivos_encontrados.append(sub_file)
        
        if not archivos_encontrados:
            return None, "No se encontraron archivos JSON o ZIP"
        
        for file_info in archivos_encontrados:
            nombre = file_info['name']
            try:
                file_response = requests.get(file_info['download_url'])
                if file_response.status_code == 200:
                    contenido = file_response.content
                    
                    if nombre.lower().endswith('.zip'):
                        with zipfile.ZipFile(io.BytesIO(contenido)) as zip_file:
                            for file_name in zip_file.namelist():
                                if file_name.lower().endswith('.json'):
                                    try:
                                        with zip_file.open(file_name) as json_file:
                                            data = json.load(json_file)
                                            if 'records' in data:
                                                records.extend(data['records'])
                                    except:
                                        pass
                    else:
                        data = json.loads(contenido)
                        if 'records' in data:
                            records.extend(data['records'])
            except:
                pass
        
        if records:
            df = pd.DataFrame(records)
            return df, f"✅ {len(records)} registros cargados"
        else:
            return None, "No se encontraron registros en los archivos"
            
    except Exception as e:
        return None, f"Error: {str(e)}"

@st.cache_data(ttl=3600, show_spinner=False)
def procesar_datos(df):
    if df is None or df.empty:
        return None
    
    df_proc = df.copy()
    
    df_proc['fecha_publicacion_dt'] = df_proc['fecha_publicacion'].apply(convertir_fecha)
    df_proc['fecha_registro_dt'] = df_proc['fecha_registro'].apply(convertir_fecha)
    df_proc['fecha_adjudicacion_dt'] = df_proc['fecha_adjudicacion'].apply(convertir_fecha)
    
    df_proc['marca'] = df_proc['descripcion'].apply(extraer_marca)
    df_proc['categoria'] = df_proc['descripcion'].apply(extraer_categoria)
    
    df_proc['precio_float'] = pd.to_numeric(df_proc['precio'], errors='coerce')
    
    df_proc['es_activo'] = (df_proc['estado_ficha'] == 'OFERTADA') & (df_proc['estado_oferta'] == 'VIGENTE')
    
    df_proc['es_junio_2026'] = (df_proc['fecha_publicacion_dt'].dt.year == 2026) & \
                               (df_proc['fecha_publicacion_dt'].dt.month == 6)
    
    df_proc['año_publicacion'] = df_proc['fecha_publicacion_dt'].dt.year
    df_proc['mes_publicacion'] = df_proc['fecha_publicacion_dt'].dt.month
    df_proc['fecha_completa'] = df_proc['fecha_publicacion_dt'].dt.strftime('%d/%m/%Y')
    
    return df_proc

def calcular_estadisticas(df):
    if df is None or df.empty:
        return None
    
    stats = {
        'total': len(df),
        'por_marca': df['marca'].value_counts().to_dict(),
        'por_categoria': df['categoria'].value_counts().to_dict(),
        'por_estado': df['estado_ficha'].value_counts().to_dict(),
        'activos': df['es_activo'].sum(),
        'precio_promedio': df['precio_float'].mean(),
        'precio_min': df['precio_float'].min(),
        'precio_max': df['precio_float'].max(),
        'precio_mediana': df['precio_float'].median(),
    }
    
    return stats

# ============ FUNCIONES DE UI ============
def mostrar_filtros(df):
    st.sidebar.markdown("---")
    st.sidebar.header("🔍 Filtros")
    df_filtrado = df.copy()
    
    categorias_opciones = ['Todas'] + sorted(df['categoria'].unique().tolist())
    categoria_seleccionada = st.sidebar.selectbox("Categoria:", categorias_opciones, key="filtro_categoria")
    if categoria_seleccionada != 'Todas':
        df_filtrado = df_filtrado[df_filtrado['categoria'] == categoria_seleccionada]
    
    marcas_opciones = ['Todas'] + sorted(df['marca'].unique().tolist())
    marca_seleccionada = st.sidebar.selectbox("Marca:", marcas_opciones, key="filtro_marca")
    if marca_seleccionada != 'Todas':
        df_filtrado = df_filtrado[df_filtrado['marca'] == marca_seleccionada]
    
    estados_opciones = ['Todos', '✅ Activos'] + sorted(df['estado_ficha'].unique().tolist())
    estado_seleccionado = st.sidebar.selectbox("Estado:", estados_opciones, key="filtro_estado")
    if estado_seleccionado == '✅ Activos':
        df_filtrado = df_filtrado[df_filtrado['es_activo'] == True]
    elif estado_seleccionado != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['estado_ficha'] == estado_seleccionado]
    
    st.sidebar.markdown("---")
    st.sidebar.markdown(f"**Mostrados:** {len(df_filtrado)} de {len(df)}")
    return df_filtrado

def mostrar_metricas(stats):
    if stats is None:
        return
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{stats['total']}</div>
            <div class="metric-label">📦 Total</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(stats['por_marca'])}</div>
            <div class="metric-label">🏷️ Marcas</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(stats['por_categoria'])}</div>
            <div class="metric-label">📂 Categorías</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">${stats['precio_promedio']:,.2f}</div>
            <div class="metric-label">💰 Precio Promedio</div>
        </div>
        """, unsafe_allow_html=True)

def mostrar_tabla(df, titulo):
    st.subheader(titulo)
    columnas = ['ID_ProductoOfertado', 'descripcion', 'marca', 'categoria', 
                'precio', 'estado_ficha', 'estado_oferta', 'fecha_publicacion']
    df_show = df[columnas].copy()
    df_show['descripcion'] = df_show['descripcion'].str[:100] + '...'
    df_show.columns = ['ID', 'Descripción', 'Marca', 'Categoría', 
                       'Precio (USD)', 'Estado Ficha', 'Estado Oferta', 'Fecha']
    st.dataframe(df_show, use_container_width=True, height=400)

def mostrar_graficos_categoria(stats):
    if stats is None:
        return
    
    col1, col2 = st.columns(2)
    
    with col1:
        if stats['por_marca']:
            df_marcas = pd.DataFrame(list(stats['por_marca'].items()), 
                                     columns=['Marca', 'Cantidad'])
            df_marcas = df_marcas.sort_values('Cantidad', ascending=False)
            fig = px.pie(df_marcas.head(10), values='Cantidad', names='Marca',
                        title='Top 10 Marcas',
                        color_discrete_sequence=px.colors.qualitative.Set3)
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(height=400, margin=dict(l=20, r=20, t=40, b=20))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
    
    with col2:
        if stats['por_categoria']:
            df_categorias = pd.DataFrame(list(stats['por_categoria'].items()), 
                                         columns=['Categoría', 'Cantidad'])
            df_categorias = df_categorias.sort_values('Cantidad', ascending=False)
            fig = px.bar(df_categorias, x='Categoría', y='Cantidad',
                        title='Productos por Categoría',
                        color='Categoría',
                        color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_traces(textposition='outside')
            fig.update_layout(height=400, margin=dict(l=20, r=20, t=40, b=20))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

def mostrar_analisis_fechas_optimizado(df):
    st.subheader("📅 Análisis por Fechas")
    
    if df.empty:
        st.warning("No hay datos para mostrar")
        return
    
    df_fechas = df.dropna(subset=['fecha_publicacion_dt']).copy()
    if df_fechas.empty:
        st.warning("No hay fechas válidas")
        return
    
    df_fechas['año'] = df_fechas['fecha_publicacion_dt'].dt.year
    df_fechas['mes'] = df_fechas['fecha_publicacion_dt'].dt.month
    df_fechas['mes_nombre'] = df_fechas['fecha_publicacion_dt'].dt.strftime('%B')
    
    col1, col2 = st.columns(2)
    
    with col1:
        años = ['Todos'] + sorted(df_fechas['año'].unique())
        año_sel = st.selectbox("Año:", años, key="fecha_año")
    
    with col2:
        meses_nombres = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                         7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
        meses_opciones = ['Todos'] + [meses_nombres[m] for m in sorted(df_fechas['mes'].unique())]
        mes_sel = st.selectbox("Mes:", meses_opciones, key="fecha_mes")
    
    if año_sel != 'Todos':
        df_fechas = df_fechas[df_fechas['año'] == año_sel]
    if mes_sel != 'Todos':
        mes_num = {v: k for k, v in meses_nombres.items()}[mes_sel]
        df_fechas = df_fechas[df_fechas['mes'] == mes_num]
    
    if df_fechas.empty:
        st.warning("No hay datos con los filtros seleccionados")
        return
    
    st.info(f"Mostrando {len(df_fechas)} registros")
    
    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Total", len(df_fechas))
    with c2:
        if not df_fechas.empty:
            st.metric("Desde", df_fechas['fecha_publicacion_dt'].min().strftime('%d/%m/%Y'))
    with c3:
        if not df_fechas.empty:
            st.metric("Hasta", df_fechas['fecha_publicacion_dt'].max().strftime('%d/%m/%Y'))
    
    if año_sel != 'Todos':
        df_meses = df_fechas.groupby('mes_nombre').size().reset_index(name='Cantidad')
        fig = px.bar(df_meses, x='mes_nombre', y='Cantidad', 
                    title=f'Productos por Mes - {año_sel}',
                    color='mes_nombre')
        fig.update_layout(height=350, showlegend=False, margin=dict(l=20, r=20, t=40, b=20))
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
    else:
        df_anios = df_fechas.groupby('año').size().reset_index(name='Cantidad')
        fig = px.bar(df_anios, x='año', y='Cantidad', 
                    title='Productos por Año',
                    color='año')
        fig.update_layout(height=350, showlegend=False, margin=dict(l=20, r=20, t=40, b=20))
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

def mostrar_buscador(df):
    st.subheader("🔍 Buscador de Número de Parte")
    st.markdown("""
    <div class="search-container">
        <p>Ingresa un número de parte para buscar en todas las descripciones de productos.</p>
        <p style="font-size: 0.9rem; color: #666;">
            <strong>Ejemplos:</strong> MDP2605F5ZZC2, M70SG6U743162000-OH, NEO55SR716100
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        numero_parte = st.text_input(
            "Número de Parte:",
            placeholder="Ej: MDP2605F5ZZC2, M70SG6U743162000-OH",
            key="buscador_input"
        )
    
    with col2:
        umbral = st.slider(
            "Umbral (%)",
            min_value=20,
            max_value=100,
            value=40,
            step=5,
            help="Porcentaje mínimo de coincidencia"
        )
    
    col1, col2 = st.columns([1, 4])
    with col1:
        buscar_click = st.button("🔍 Buscar", use_container_width=True, type="primary")
    
    if 'resultados_busqueda' not in st.session_state:
        st.session_state.resultados_busqueda = None
    if 'ultima_busqueda' not in st.session_state:
        st.session_state.ultima_busqueda = ""
    
    if buscar_click and numero_parte:
        with st.spinner(f"Buscando '{numero_parte}'..."):
            resultados_df = buscar_numero_parte(df, numero_parte, umbral)
            st.session_state.resultados_busqueda = resultados_df
            st.session_state.ultima_busqueda = numero_parte
    
    if st.session_state.resultados_busqueda is not None:
        resultados_df = st.session_state.resultados_busqueda
        
        if not resultados_df.empty:
            st.markdown(f"""
            <div class="result-header">
                <strong>🔍 Resultados para: {st.session_state.ultima_busqueda}</strong><br>
                Se encontraron <strong>{len(resultados_df)}</strong> coincidencias
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total", len(resultados_df))
            with col2:
                activos = resultados_df[resultados_df['es_activo'] == True].shape[0]
                st.metric("✅ Activos", activos)
            with col3:
                junio = resultados_df[resultados_df['es_junio_2026'] == True].shape[0]
                st.metric("📅 Junio 2026", junio)
            with col4:
                max_coinc = resultados_df['coincidencia'].max()
                st.metric("🎯 Mejor coincidencia", f"{max_coinc}%")
            
            st.divider()
            
            tabla_show = resultados_df[[
                'ID', 'marca', 'categoria', 'precio', 
                'estado_ficha', 'estado_oferta', 'fecha_publicacion', 
                'coincidencia', 'es_activo', 'es_junio_2026'
            ]].copy()
            
            tabla_show['coincidencia'] = tabla_show['coincidencia'].apply(lambda x: f"{x}%")
            tabla_show['precio'] = tabla_show['precio'].apply(lambda x: f"${float(x):,.2f}" if x else "$0")
            tabla_show['estado'] = tabla_show['es_activo'].apply(
                lambda x: '🟢 Activo' if x else '🔴 Inactivo'
            )
            tabla_show['junio'] = tabla_show['es_junio_2026'].apply(
                lambda x: '✅ Junio 2026' if x else ''
            )
            tabla_show['descripcion'] = resultados_df['descripcion'].str[:150] + '...'
            
            tabla_show = tabla_show[[
                'ID', 'descripcion', 'marca', 'categoria', 'precio', 
                'estado', 'estado_ficha', 'estado_oferta', 
                'fecha_publicacion', 'coincidencia', 'junio'
            ]]
            
            tabla_show.columns = [
                'ID', 'Descripción', 'Marca', 'Categoría', 'Precio (USD)',
                'Estado', 'Estado Ficha', 'Estado Oferta', 
                'Fecha Publicación', 'Coincidencia', 'Junio 2026'
            ]
            
            st.dataframe(tabla_show, use_container_width=True, height=500)
            
            st.subheader("💾 Exportar Resultados")
            col1, col2 = st.columns(2)
            
            with col1:
                csv = resultados_df.to_csv(index=False)
                b64 = base64.b64encode(csv.encode()).decode()
                href = f'<a href="data:file/csv;base64,{b64}" download="busqueda_{st.session_state.ultima_busqueda}.csv" style="background-color: #1f77b4; color: white; padding: 8px 16px; border-radius: 5px; text-decoration: none;">📥 Descargar CSV</a>'
                st.markdown(href, unsafe_allow_html=True)
            
            with col2:
                json_str = resultados_df.to_json(orient='records', date_format='iso')
                b64_json = base64.b64encode(json_str.encode()).decode()
                href_json = f'<a href="data:file/json;base64,{b64_json}" download="busqueda_{st.session_state.ultima_busqueda}.json" style="background-color: #28a745; color: white; padding: 8px 16px; border-radius: 5px; text-decoration: none;">📥 Descargar JSON</a>'
                st.markdown(href_json, unsafe_allow_html=True)
                
        else:
            st.markdown(f"""
            <div class="no-result">
                <strong>🔍 No se encontraron resultados para '{st.session_state.ultima_busqueda}'</strong><br>
                Prueba con otro número de parte o reduce el umbral de coincidencia.
            </div>
            """, unsafe_allow_html=True)

# ============ RESUMEN JUNIO 2026 ACTIVO ============
def mostrar_resumen_junio_activo(df):
    """Muestra SOLO las fichas de Junio 2026 con estado OFERTADA y VIGENTE y permite descargar Excel formateado"""
    
    # FILTRO ESTRICTO
    df_filtrado = df[
        (df['es_junio_2026'] == True) & 
        (df['estado_ficha'] == 'OFERTADA') & 
        (df['estado_oferta'] == 'VIGENTE')
    ]
    
    if df_filtrado.empty:
        st.warning("⚠️ No hay fichas en Junio 2026 con estado OFERTADA y VIGENTE")
        return
    
    stats = calcular_estadisticas(df_filtrado)
    
    # ============ HEADER ============
    st.markdown("""
    <div class="resumen-activo-header">
        <h2>✅ JUNIO 2026 - FICHAS ACTIVAS</h2>
        <p>Fichas con estado OFERTADA y VIGENTE</p>
    </div>
    """, unsafe_allow_html=True)
    
    # ============ MÉTRICAS ============
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card" style="background: linear-gradient(135deg, #28a745 0%, #20c997 100%);">
            <div class="metric-value" style="color: white;">{stats['total']}</div>
            <div class="metric-label" style="color: white;">📦 Fichas Activas</div>
            <div class="metric-sub" style="color: rgba(255,255,255,0.8);">Junio 2026</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(stats['por_marca'])}</div>
            <div class="metric-label">🏷️ Marcas</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(stats['por_categoria'])}</div>
            <div class="metric-label">📂 Categorías</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">${stats['precio_promedio']:,.2f}</div>
            <div class="metric-label">💰 Precio Promedio</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col5:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">${stats['precio_max']:,.2f}</div>
            <div class="metric-label">💰 Precio Máximo</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.divider()
    
    # ============ GRÁFICOS ============
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🏷️ Marcas")
        if stats['por_marca']:
            df_marcas = pd.DataFrame(list(stats['por_marca'].items()), 
                                     columns=['Marca', 'Cantidad'])
            df_marcas = df_marcas.sort_values('Cantidad', ascending=False)
            
            fig = px.bar(df_marcas.head(15), x='Marca', y='Cantidad',
                        title=f'Top 15 Marcas (Total: {stats["total"]} activas)',
                        color='Marca', text='Cantidad',
                        color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_traces(textposition='outside')
            fig.update_layout(height=400, showlegend=False, margin=dict(l=20, r=20, t=40, b=40))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
    
    with col2:
        st.subheader("📂 Categorías")
        if stats['por_categoria']:
            df_cat = pd.DataFrame(list(stats['por_categoria'].items()), 
                                  columns=['Categoría', 'Cantidad'])
            df_cat = df_cat.sort_values('Cantidad', ascending=False)
            
            fig = px.pie(df_cat, values='Cantidad', names='Categoría',
                        title='Distribución por Categoría',
                        color_discrete_sequence=px.colors.qualitative.Set3)
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(height=400, margin=dict(l=20, r=20, t=40, b=20))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
    
    st.divider()
    
    # ============ TABLA DE DETALLE ============
    st.subheader("📋 Detalle de Fichas Activas")
    
    columnas_mostrar = ['ID_ProductoOfertado', 'descripcion', 'marca', 'categoria', 
                        'precio', 'estado_ficha', 'estado_oferta', 'fecha_publicacion']
    
    df_show = df_filtrado[columnas_mostrar].copy()
    df_show['descripcion'] = df_show['descripcion'].str[:150] + '...'
    df_show.columns = ['ID', 'Descripción', 'Marca', 'Categoría', 
                       'Precio (USD)', 'Estado Ficha', 'Estado Oferta', 'Fecha Publicación']
    
    st.dataframe(df_show, use_container_width=True, height=400)
    
    st.divider()
    
    # ============ BOTÓN PARA DESCARGAR EXCEL CON FORMATO ============
    st.subheader("📥 Descargar Reporte Excel Formateado")
    
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        # Excel con formato
        excel_file = crear_excel_formateado(df_filtrado, "junio_2026_activas.xlsx")
        b64 = base64.b64encode(excel_file.getvalue()).decode()
        href = f'''
        <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" 
           download="junio_2026_activas.xlsx" 
           style="background-color: #28a745; color: white; padding: 10px 20px; border-radius: 5px; text-decoration: none; display: inline-block; font-weight: bold;">
            📊 Descargar Excel (Con Formato)
        </a>
        '''
        st.markdown(href, unsafe_allow_html=True)
    
    with col2:
        # CSV simple
        csv = df_filtrado.to_csv(index=False)
        b64_csv = base64.b64encode(csv.encode()).decode()
        href_csv = f'''
        <a href="data:file/csv;base64,{b64_csv}" 
           download="junio_2026_activas.csv" 
           style="background-color: #1f77b4; color: white; padding: 10px 20px; border-radius: 5px; text-decoration: none; display: inline-block; font-weight: bold;">
            📥 Descargar CSV
        </a>
        '''
        st.markdown(href_csv, unsafe_allow_html=True)
    
    with col3:
        # JSON
        json_str = df_filtrado.to_json(orient='records', date_format='iso')
        b64_json = base64.b64encode(json_str.encode()).decode()
        href_json = f'''
        <a href="data:file/json;base64,{b64_json}" 
           download="junio_2026_activas.json" 
           style="background-color: #6c757d; color: white; padding: 10px 20px; border-radius: 5px; text-decoration: none; display: inline-block; font-weight: bold;">
            📥 Descargar JSON
        </a>
        '''
        st.markdown(href_json, unsafe_allow_html=True)
    
    # Leyenda de colores
    with st.expander("🎨 Leyenda de Colores del Excel"):
        st.markdown("""
        | Color | Significado |
        |-------|-------------|
        | 🟢 **Verde** | OFERTADA + VIGENTE (Activo) |
        | 🟡 **Amarillo** | OFERTADA + Otro estado |
        | 🔴 **Rojo** | EXCLUIDA |
        | ⬜ **Gris** | Otros estados |
        | 🔵 **Azul** | Encabezados |
        """)
        
        st.info("""
        **El Excel incluye:**
        - ✅ Colores por estado (verde para activas)
        - ✅ Formato de precios en USD
        - ✅ Bordes y alineación profesional
        - ✅ Hoja de Resumen con estadísticas
        - ✅ Hojas de Marcas y Categorías
        - ✅ Congelación de panel para fácil navegación
        """)

# ============ MAIN ============
def main():
    st.markdown('<h1 class="main-header">📊 Dashboard de Análisis de Productos</h1>', unsafe_allow_html=True)
    
    if 'datos_cargados' not in st.session_state:
        st.session_state.datos_cargados = False
    
    with st.sidebar:
        st.header("⚙️ Configuración")
        
        repo_owner = st.text_input("Owner:", value="StalinHA", key="repo_owner")
        repo_name = st.text_input("Repositorio:", value="fichadjudicados", key="repo_name")
        folder = st.text_input("Carpeta (opcional):", value="", key="repo_folder")
        
        if st.button("🚀 Cargar Datos", use_container_width=True, type="primary"):
            with st.spinner("Cargando..."):
                df, mensaje = cargar_datos_github(repo_owner, repo_name, "main", folder)
                if df is not None:
                    st.session_state.df_raw = df
                    st.session_state.df_proc = procesar_datos(df)
                    st.session_state.datos_cargados = True
                    st.session_state.mensaje_carga = mensaje
                    st.rerun()
                else:
                    st.error(mensaje)
        
        if st.session_state.datos_cargados:
            st.success(st.session_state.mensaje_carga)
            st.info(f"Registros: {len(st.session_state.df_proc)}")
    
    if not st.session_state.datos_cargados:
        st.info("📁 Haz clic en 'Cargar Datos' en el panel izquierdo")
        return
    
    df = st.session_state.df_proc
    df_filtrado = mostrar_filtros(df)
    
    if df_filtrado.empty:
        st.warning("No hay datos con los filtros seleccionados")
        return
    
    stats = calcular_estadisticas(df_filtrado)
    
    # ============ APARTADO: RESUMEN JUNIO 2026 ACTIVO ============
    mostrar_resumen_junio_activo(df)
    
    st.divider()
    
    # ============ JUNIO 2026 (TODAS) ============
    df_junio = df[df['es_junio_2026'] == True]
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 1rem; border-radius: 10px; margin: 1rem 0; text-align: center;">
        <h2 style="color: white; margin: 0;">📊 JUNIO 2026 - TODAS LAS FICHAS</h2>
        <p style="color: rgba(255,255,255,0.8); margin: 0;">Incluye todos los estados</p>
    </div>
    """, unsafe_allow_html=True)
    
    if not df_junio.empty:
        stats_junio = calcular_estadisticas(df_junio)
        
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{stats_junio['total']}</div>
                <div class="metric-label">📦 Total Junio</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{len(stats_junio['por_marca'])}</div>
                <div class="metric-label">🏷️ Marcas</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">{len(stats_junio['por_categoria'])}</div>
                <div class="metric-label">📂 Categorías</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-value">${stats_junio['precio_promedio']:,.2f}</div>
                <div class="metric-label">💰 Promedio</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col5:
            st.markdown(f"""
            <div class="metric-card" style="background: #28a745; color: white;">
                <div class="metric-value" style="color: white;">{stats_junio['activos']}</div>
                <div class="metric-label" style="color: white;">✅ Activas</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.divider()
        
        col1, col2 = st.columns(2)
        with col1:
            if stats_junio['por_marca']:
                df_marcas = pd.DataFrame(list(stats_junio['por_marca'].items()), 
                                         columns=['Marca', 'Cantidad'])
                df_marcas = df_marcas.sort_values('Cantidad', ascending=False)
                fig = px.bar(df_marcas.head(10), x='Marca', y='Cantidad',
                            title='Top 10 Marcas',
                            color='Marca', text='Cantidad')
                fig.update_traces(textposition='outside')
                fig.update_layout(height=350, showlegend=False, margin=dict(l=20, r=20, t=40, b=20))
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        
        with col2:
            if stats_junio['por_categoria']:
                df_cat = pd.DataFrame(list(stats_junio['por_categoria'].items()), 
                                      columns=['Categoría', 'Cantidad'])
                fig = px.pie(df_cat, values='Cantidad', names='Categoría',
                            title='Categorías',
                            color_discrete_sequence=px.colors.qualitative.Set2)
                fig.update_traces(textposition='inside', textinfo='percent+label')
                fig.update_layout(height=350, margin=dict(l=20, r=20, t=40, b=20))
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        
        st.divider()
        mostrar_tabla(df_junio, "📋 Detalle - Junio 2026")
    
    else:
        st.warning("No se encontraron fichas en Junio 2026")
    
    st.divider()
    
    # ============ PESTAÑAS PRINCIPALES ============
    st.markdown("""
    <div class="analisis-header">
        <h2>📊 ANÁLISIS COMPLETO</h2>
    </div>
    """, unsafe_allow_html=True)
    
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Resumen", "🏷️ Marcas", "📂 Categorías", "📅 Fechas", "🔍 Buscador", "📋 Datos"
    ])
    
    with tab1:
        mostrar_metricas(stats)
        st.divider()
        mostrar_graficos_categoria(stats)
    
    with tab2:
        st.subheader("🏷️ Análisis por Marca")
        if stats['por_marca']:
            df_marcas = pd.DataFrame(list(stats['por_marca'].items()), 
                                     columns=['Marca', 'Cantidad'])
            df_marcas = df_marcas.sort_values('Cantidad', ascending=False)
            fig = px.bar(df_marcas, x='Marca', y='Cantidad', 
                        color='Marca', text='Cantidad')
            fig.update_traces(textposition='outside')
            fig.update_layout(height=400, showlegend=False, margin=dict(l=20, r=20, t=40, b=20))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
            st.dataframe(df_marcas, use_container_width=True)
    
    with tab3:
        st.subheader("📂 Análisis por Categoría")
        if stats['por_categoria']:
            df_cat = pd.DataFrame(list(stats['por_categoria'].items()), 
                                  columns=['Categoría', 'Cantidad'])
            df_cat = df_cat.sort_values('Cantidad', ascending=False)
            fig = px.pie(df_cat, values='Cantidad', names='Categoría',
                        color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(height=400, margin=dict(l=20, r=20, t=40, b=20))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
            st.dataframe(df_cat, use_container_width=True)
    
    with tab4:
        mostrar_analisis_fechas_optimizado(df_filtrado)
    
    with tab5:
        mostrar_buscador(df_filtrado)
    
    with tab6:
        mostrar_tabla(df_filtrado, "📋 Todos los Productos")
        
        st.subheader("💾 Exportar")
        col1, col2 = st.columns(2)
        
        with col1:
            csv = df_filtrado.to_csv(index=False)
            b64 = base64.b64encode(csv.encode()).decode()
            href = f'<a href="data:file/csv;base64,{b64}" download="productos.csv" style="background-color: #1f77b4; color: white; padding: 8px 16px; border-radius: 5px; text-decoration: none;">📥 CSV</a>'
            st.markdown(href, unsafe_allow_html=True)
        
        with col2:
            json_str = df_filtrado.to_json(orient='records', date_format='iso')
            b64_json = base64.b64encode(json_str.encode()).decode()
            href_json = f'<a href="data:file/json;base64,{b64_json}" download="productos.json" style="background-color: #28a745; color: white; padding: 8px 16px; border-radius: 5px; text-decoration: none;">📥 JSON</a>'
            st.markdown(href_json, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
